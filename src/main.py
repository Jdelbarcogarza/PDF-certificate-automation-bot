
from fpdf import FPDF
from openpyxl import load_workbook
from time import sleep
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication



def main():
    
    ### Get information from participants with Openpyxl

    workbook = load_workbook(filename='registro.xlsx')

    # get  active worksheet
    sheet = workbook['Respuestas de formulario 1']

    # personal credentials to send emails
    username = "USERNAME TO LOG IN"
    password = "PERSONAL PASSWORD"
    mail_from = "EMAIL FROM SENDER"
    mail_subject = "CERTIFICADO DE PARTICIPACIÓN"
    mail_body = "MAIL BODY"

    path_to_pdf = 'WHERE IS THE GENERATED PDF STORED'

    certificateID = 0

    for val in sheet.iter_rows(max_col=4):

        # if student attended event
        if val[0].value == 1:

            # orientation Landscape
            pdf = FPDF(orientation='L')

            # add page that will serve as canvas
            pdf.add_page()

            # add image to PDF
            pdf.image('CERTIFICADO.png', w=300, h=210, x=0.0, y=0.0)

            # set font to name
            pdf.set_font('Arial','B', 24)

            # place cursor in location where the name goes
            pdf.set_xy(10, 128)

            # function that prints the names.
            pdf.cell(280,10,val[3].value, align='C')

            pdf.output(path_to_pdf+str(certificateID)+'.pdf', 'F')

            mimemsg = MIMEMultipart()
            mimemsg['From']=mail_from
            mimemsg['To']=val[2].value
            mimemsg['Subject']=mail_subject
            mimemsg.attach(MIMEText(mail_body, 'plain'))

            with open(path_to_pdf+str(certificateID)+'.pdf', 'rb') as f:
                attach = MIMEApplication(f.read(), _subtype='pdf')

            attach.add_header('Content-Disposition','attachment',filename="[NAME OF FILE ATTACHMENT].pdf" )
            mimemsg.attach(attach)

            connection = smtplib.SMTP(host='smtp.office365.com', port=587)
            connection.starttls()
            connection.login(username,password)
            connection.send_message(mimemsg)
            connection.quit()    

            certificateID = certificateID + 1

            # order of output: attended, email, full name
            print('Mail sent to:', val[3].value, 'at email:', val[2].value)
            sleep(1)
            


    



    



if __name__ == '__main__':
    main()